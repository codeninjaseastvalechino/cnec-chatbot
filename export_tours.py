"""
Export GBS tours and student appointments to Excel format.

Usage:
    from export_tours import create_excel_file, create_unified_excel_file
    excel_path = await create_excel_file(sessions)
    excel_path = await create_unified_excel_file(gbs_sessions, appointments)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from typing import List, Union, Dict, Optional, Any
import asyncio


async def create_excel_file(sessions, filename=None):
    """
    Create an Excel file with GBS tours.

    Args:
        sessions: List of GBSSession objects
        filename: Optional custom filename (default: tours_YYYY-MM-DD.xlsx)

    Returns:
        Path to the created Excel file
    """
    if not sessions:
        raise ValueError("No tours to export")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tours"

    # Define styles
    header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths (Time, Parent, Children, Tour Type, Assigned To)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    # Add headers
    headers = ['Time', 'Parent', 'Children', 'Tour Type', 'Assigned To']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Add data rows
    for row_num, session in enumerate(sessions, 2):
        # Time
        time_cell = ws.cell(row=row_num, column=1)
        time_cell.value = session.time_display()
        time_cell.border = border
        time_cell.alignment = Alignment(horizontal="center")

        # Student (Guardian)
        student_cell = ws.cell(row=row_num, column=2)
        student_cell.value = session.student_name
        student_cell.border = border

        # Children
        children_cell = ws.cell(row=row_num, column=3)
        children_text = ", ".join(session.child_display) if session.child_display else "N/A"
        children_cell.value = children_text
        children_cell.border = border

        # Tour Type
        tour_type_cell = ws.cell(row=row_num, column=4)
        tour_type_cell.value = session.tour_type
        tour_type_cell.border = border
        tour_type_cell.alignment = Alignment(horizontal="center")

        # Assigned To
        assigned_cell = ws.cell(row=row_num, column=5)
        assigned_cell.value = session.assignee_name
        assigned_cell.border = border

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Add summary at bottom
    summary_row = len(sessions) + 3
    ws.cell(row=summary_row, column=1).value = f"Total Tours: {len(sessions)}"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    gbs_count = sum(1 for s in sessions if s.tour_type == "GBS")
    jrgbs_count = sum(1 for s in sessions if s.tour_type == "JR GBS")
    ws.cell(row=summary_row + 1, column=1).value = f"GBS: {gbs_count} | JR GBS: {jrgbs_count}"

    # Save file
    if not filename:
        date_str = sessions[0].date_display() if sessions else datetime.now().strftime("%Y-%m-%d")
        filename = f"tours_{date_str.replace(' ', '_')}.xlsx"

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)

    filepath = export_dir / filename
    wb.save(filepath)

    return str(filepath)


def create_camps_excel_file(camps, rosters=None, filename=None):
    """
    Create an Excel file for camp data.

    If rosters contains kids for a camp, writes a per-kid roster sheet.
    Falls back to a summary sheet (one row per camp) when no roster data.

    Args:
        camps: List of CampRecord objects
        rosters: Dict of event_id -> List[CampKid] (or None / empty)
        filename: Optional custom filename

    Returns:
        Path to the created Excel file
    """
    if not camps:
        raise ValueError("No camp data to export")

    rosters = rosters or {}

    wb = Workbook()

    header_fill = PatternFill(start_color="162044", end_color="162044", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="1a2a4a", end_color="1a2a4a", fill_type="solid")
    subheader_font = Font(bold=True, color="eabb5c", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Determine whether we have any roster data
    has_roster = any(rosters.get(c.event_id) is not None for c in camps)

    if has_roster:
        # Roster sheet: one row per kid, grouped by camp
        ws = wb.active
        ws.title = "Camp Roster"

        ws.column_dimensions["A"].width = 30   # Camp
        ws.column_dimensions["B"].width = 14   # Date
        ws.column_dimensions["C"].width = 20   # Time
        ws.column_dimensions["D"].width = 22   # Kid Name
        ws.column_dimensions["E"].width = 8    # Age
        ws.column_dimensions["F"].width = 22   # Parent
        ws.column_dimensions["G"].width = 15   # Phone
        ws.column_dimensions["H"].width = 12   # Status

        headers = ["Camp", "Date", "Time", "Kid Name", "Age", "Parent", "Phone", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 22

        row = 2
        for camp in camps:
            kids = rosters.get(camp.event_id)
            if kids is None:
                # No roster available — write a single placeholder row
                ws.cell(row=row, column=1, value=camp.title).border = border
                ws.cell(row=row, column=2, value=camp.start_dt.strftime("%b %-d")).border = border
                ws.cell(row=row, column=3, value=camp.time_range()).border = border
                ws.cell(row=row, column=4, value="(roster unavailable)").border = border
                for col in range(5, 9):
                    ws.cell(row=row, column=col).border = border
                row += 1
                continue

            if not kids:
                ws.cell(row=row, column=1, value=camp.title).border = border
                ws.cell(row=row, column=2, value=camp.start_dt.strftime("%b %-d")).border = border
                ws.cell(row=row, column=3, value=camp.time_range()).border = border
                ws.cell(row=row, column=4, value="(no enrollments)").border = border
                for col in range(5, 9):
                    ws.cell(row=row, column=col).border = border
                row += 1
                continue

            for kid in kids:
                ws.cell(row=row, column=1, value=camp.title).border = border
                ws.cell(row=row, column=2, value=camp.start_dt.strftime("%b %-d")).border = border
                ws.cell(row=row, column=3, value=camp.time_range()).border = border
                ws.cell(row=row, column=4, value=kid.participant_name).border = border
                ws.cell(row=row, column=5, value=kid.age or "").border = border
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=6, value=kid.buyer_name).border = border
                ws.cell(row=row, column=7, value=kid.phone).border = border
                ws.cell(row=row, column=8, value=kid.status).border = border
                row += 1

    else:
        # Summary sheet: one row per camp
        ws = wb.active
        ws.title = "Camps"

        ws.column_dimensions["A"].width = 40   # Camp name
        ws.column_dimensions["B"].width = 14   # Date
        ws.column_dimensions["C"].width = 20   # Time
        ws.column_dimensions["D"].width = 12   # Enrolled
        ws.column_dimensions["E"].width = 12   # Capacity
        ws.column_dimensions["F"].width = 12   # Spots Left

        headers = ["Camp", "Date", "Time", "Enrolled", "Capacity", "Spots Left"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 22

        for row_num, camp in enumerate(camps, 2):
            spots = camp.spots_left()
            ws.cell(row=row_num, column=1, value=camp.title).border = border
            ws.cell(row=row_num, column=2, value=camp.start_dt.strftime("%b %-d, %Y")).border = border
            ws.cell(row=row_num, column=3, value=camp.time_range()).border = border
            ws.cell(row=row_num, column=4, value=camp.enrolled).border = border
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=5, value=camp.capacity if camp.capacity else "").border = border
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=6, value=spots if spots is not None else "").border = border
            ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center")

    # Summary row at bottom
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value=f"Total camps: {len(camps)}").font = Font(bold=True)
    if has_roster:
        total_kids = sum(
            len([k for k in (rosters.get(c.event_id) or []) if k.status.lower() not in ("cancelled",)])
            for c in camps
        )
        ws.cell(row=last_row + 1, column=1, value=f"Total active enrollments: {total_kids}").font = Font(bold=True)

    if not filename:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"camps_{date_str}.xlsx"

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)

    filepath = export_dir / filename
    wb.save(filepath)
    return str(filepath)


async def create_unified_excel_file(gbs_sessions, appointments, filename=None):
    """
    Create an Excel file with unified schedule (GBS tours + student appointments).

    Creates a single merged, time-ordered schedule with columns:
    Time | Student | Type | Duration | Instructor/Staff | Location | Notes

    Args:
        gbs_sessions: List of GBSSession objects
        appointments: List of StudentAppointment objects
        filename: Optional custom filename (default: schedule_YYYY-MM-DD.xlsx)

    Returns:
        Path to the created Excel file
    """
    if not gbs_sessions and not appointments:
        raise ValueError("No sessions or appointments to export")

    # Merge and sort by time
    items = []

    for session in gbs_sessions:
        # For GBS: student = child name with age, parent = guardian name
        child_names = ", ".join(session.child_display) if session.child_display else ""
        items.append({
            "time": session.start_time,
            "student": child_names,
            "type": session.tour_type,
            "belt": "",
            "parent": session.student_name,  # guardian
        })

    for appt in appointments:
        items.append({
            "time": appt.start_time,
            "student": appt.student_name,
            "type": appt.appointment_type,
            "belt": appt.rank,
            "parent": getattr(appt, "parent_name", "") or "",
        })

    # Sort by time — strip timezone info so naive and aware datetimes can be compared
    items.sort(key=lambda x: x["time"].astimezone().replace(tzinfo=None) if x["time"].tzinfo else x["time"])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Define styles
    header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths: Time | Student | Type | Belt | Parent
    ws.column_dimensions['A'].width = 12  # Time
    ws.column_dimensions['B'].width = 22  # Student
    ws.column_dimensions['C'].width = 18  # Type
    ws.column_dimensions['D'].width = 15  # Belt
    ws.column_dimensions['E'].width = 22  # Parent

    # Add headers
    headers = ['Time', 'Student', 'Type', 'Belt', 'Parent']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Add data rows
    for row_num, item in enumerate(items, 2):
        t = item["time"]
        if t.tzinfo is not None:
            t = t.astimezone()  # Convert UTC → local for GBS sessions

        # Time
        ws.cell(row=row_num, column=1).value = t.strftime("%I:%M %p").lstrip("0")
        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

        # Student
        ws.cell(row=row_num, column=2).value = item["student"]
        ws.cell(row=row_num, column=2).border = border

        # Type
        ws.cell(row=row_num, column=3).value = item["type"]
        ws.cell(row=row_num, column=3).border = border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")

        # Belt
        ws.cell(row=row_num, column=4).value = item["belt"]
        ws.cell(row=row_num, column=4).border = border
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")

        # Parent
        ws.cell(row=row_num, column=5).value = item["parent"]
        ws.cell(row=row_num, column=5).border = border

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Add summary at bottom
    summary_row = len(items) + 3
    ws.cell(row=summary_row, column=1).value = f"Total: {len(items)} items"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    gbs_count = sum(1 for s in gbs_sessions)
    appt_count = sum(1 for a in appointments)
    ws.cell(row=summary_row + 1, column=1).value = f"GBS Tours: {gbs_count} | Appointments: {appt_count}"

    # Save file
    if not filename:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"schedule_{date_str}.xlsx"

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)

    filepath = export_dir / filename
    wb.save(filepath)

    return str(filepath)
